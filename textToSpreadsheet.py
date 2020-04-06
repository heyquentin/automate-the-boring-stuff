#!/usr/bin/env python3
# Write a program to read in the contents of several text files (you can make the text files yourself) and insert those contents into a spreadsheet, with one line 
# of text per row. The lines of the first text file will be in the cells of column A, the lines of the second text file will be in the cells of column B, and so on.

from pathlib import Path
import os
import openpyxl

# TODO: create a spreadsheet to put all of the lines into
wb = openpyxl.Workbook()
sheet = wb.active

# TODO: Get all txt files from the working directory
p = Path(os.getcwd())
txtDocs = list(p.glob('*.txt')) # creates a list of the .txt documents in the folder

# TODO: set up counters for the rows and columns
colCounter = 1
rowCounter = 1

# TODO: Read each text file's contents
for txtFile in txtDocs:                     # for each text file..
    openFile = open(txtFile)                # open the file
    textContent = openFile.readlines()      # read all the lines in the file
    numOfLinesInFile = len(textContent)     # get a total number of lines in each text doc. You need this to know how many lines to write to the spreadsheet

    # TODO: put each list entry onto a new line in the spreadsheet
    for i in range(0,numOfLinesInFile):     # how ever many lines are in the text doc, that's how big this range is
        eachLine = textContent[i]           # textContent is each line of the txt file as per readlines above
        myCell = sheet.cell(row=rowCounter, column=colCounter)  # set up a variable with the sheet's cell defaults (row 1, column 1)
        myCell.value = eachLine             # put the text content into the cell value
        rowCounter += 1                     # increment the row counter so it gets a new row to put a new line into
    colCounter += 1                     # increment the column counter once we're out of the for loop above because we want to go to the next column
    rowCounter = 1                      # set the row counter back to 1 otherwise it'll start somewhere that we don't want to start at. we want to start at 1

# this is here to show that it's working without having to open in Excel
# I used 3 txt documents so that's why there's only 3 columns
maxRow = int(sheet.max_row)
for eachRow in range(1,maxRow):
    print('Column 1: ' + str(sheet.cell(row=eachRow, column=1).value))
    print('Column 2: ' + str(sheet.cell(row=eachRow, column=2).value))
    print('Column 3: ' + str(sheet.cell(row=eachRow, column=3).value))